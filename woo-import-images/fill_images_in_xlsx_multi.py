#!/usr/bin/env python3
"""
Fill the 'Image' column in ALL sheets of an Excel workbook by finding a first image URL
for each product Title (using a simple Bing Images page scrape).

Usage (inside your virtual environment):
    python3 fill_images_in_xlsx_multi.py

Prereqs:
    pip install openpyxl requests beautifulsoup4 lxml
"""

import time
import urllib.parse
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# ================== CONFIG ==================
INPUT_XLSX = "products_for_auto_images_ALL_SHEETS.xlsx"  # Adjust if you rename the file
OUTPUT_XLSX = "products_with_images_ALL_SHEETS.xlsx"
TITLE_HEADER = "Title"
IMAGE_HEADER = "Image"
REQUEST_TIMEOUT = 20
DELAY_BETWEEN_QUERIES = 1.2  # seconds - be polite
# ============================================

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
}

def find_first_image_url_bing(query: str) -> str | None:
    """
    Naive scraper for Bing Images. This may break if Bing changes markup.
    Consider an official API for production use.
    """
    q = urllib.parse.quote_plus(query)
    url = f"https://www.bing.com/images/search?q={q}&form=HDRSC2"
    try:
        r = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
    except Exception:
        return None

    from bs4 import BeautifulSoup
    soup = BeautifulSoup(r.text, "lxml")

    # Strategy 1: anchors with class iusc and a JSON-ish "m" attribute that contains "murl"
    for a in soup.select("a.iusc"):
        m = a.get("m", "")
        key = '"murl":"'
        if key in m:
            start = m.find(key) + len(key)
            end = m.find('"', start)
            if start > -1 and end > start:
                candidate = m[start:end]
                if candidate.startswith("http"):
                    return candidate

    # Strategy 2: fallback to <img> tags with http(s) src/data-src and an image extension
    for img in soup.select("img"):
        src = img.get("data-src") or img.get("src") or ""
        if src.startswith("http") and any(ext in src.lower() for ext in [".jpg", ".jpeg", ".png", ".webp"]):
            return src

    return None

def ensure_headers(ws, headers_needed):
    """
    Ensure the given headers exist in the first row.
    Return a mapping {header_name: column_index}.
    """
    header_row = 1
    existing = {}
    max_col = ws.max_column or 1

    # Read existing headers
    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if isinstance(val, str):
            existing[val.strip()] = col

    # Add any missing headers at the end
    for h in headers_needed:
        if h not in existing:
            max_col += 1
            ws.cell(row=header_row, column=max_col, value=h)
            existing[h] = max_col

    return existing

def main():
    wb = load_workbook(INPUT_XLSX)
    total_filled = 0
    total_skipped = 0

    for ws in wb.worksheets:
        # Ensure headers for each sheet
        col_map = ensure_headers(ws, [TITLE_HEADER, IMAGE_HEADER])
        title_col = col_map[TITLE_HEADER]
        image_col = col_map[IMAGE_HEADER]

        rows = ws.max_row
        filled = 0
        skipped = 0

        print(f"\nSheet: {ws.title} — rows: {rows-1}")
        print(f"Filling '{IMAGE_HEADER}' from '{TITLE_HEADER}'...")

        for row in range(2, rows + 1):
            title = ws.cell(row=row, column=title_col).value
            if not title or not str(title).strip():
                skipped += 1
                continue

            # Skip if already has an image URL
            existing_image = ws.cell(row=row, column=image_col).value
            if existing_image and str(existing_image).strip():
                continue

            title_str = str(title).strip()
            print(f"- [{ws.title} R{row}] Searching image for: {title_str!r}")
            url = find_first_image_url_bing(title_str)

            if url:
                ws.cell(row=row, column=image_col, value=url)
                filled += 1
            else:
                print("  (no image found)")

            time.sleep(DELAY_BETWEEN_QUERIES)

        print(f"Sheet '{ws.title}': filled {filled}, skipped {skipped}")
        total_filled += filled
        total_skipped += skipped

    wb.save(OUTPUT_XLSX)
    print(f"\nAll done. Total filled {total_filled}, total skipped {total_skipped}.")
    print(f"Saved: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
